# -*- coding: utf-8 -*-
import os
import sys
import argparse
import re
import csv
import subprocess
import time
import shutil
import config
import threading

# Force UTF-8 output encoding for Windows console
try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass  # Fallback for older Python versions

# --- CAU HINH DUONG DAN ---
# MT5_PATH sẽ được cập nhật trong run_auto() nếu argument được cung cấp
MT5_PATH = config.MT5_PATH
INI_BASE_DIR = config.BACKTEST_INI_DIR
BASE_RESULT_DIR = config.BASE_RESULT_DIR
DIR_XLSX = os.path.join(BASE_RESULT_DIR, "xlsx")

MT5_TERMINAL_ROOT = config.TERMINAL_DATA_PATH
MT5_TESTER_FILES = os.path.join(MT5_TERMINAL_ROOT, r"Tester\Files")
MT5_TESTER_REPORTS = os.path.join(MT5_TERMINAL_ROOT, r"Tester\Reports")
MT5_FILES = os.path.join(MT5_TERMINAL_ROOT, r"MQL5\Files")
MT5_REPORTS = os.path.join(MT5_TERMINAL_ROOT, r"reports")

SEARCH_DIRS = [MT5_TESTER_REPORTS, MT5_TESTER_FILES, MT5_FILES, MT5_REPORTS, BASE_RESULT_DIR, MT5_TERMINAL_ROOT]
TIMEOUT_SECONDS = 1800000

os.makedirs(DIR_XLSX, exist_ok=True)


def build_paths(terminal_data_path=None):
    paths = config.init_paths(terminal_data_path or config.DEFAULT_TERMINAL_DATA_PATH)
    config.TERMINAL_DATA_PATH = paths["TERMINAL_DATA_PATH"]
    config.INI_SOURCE_DIR = paths["INI_SOURCE_DIR"]
    config.OPTIMIZE_XML_DIR = paths["OPTIMIZE_XML_DIR"]
    config.FILTERED_CSV_DIR = paths["FILTERED_CSV_DIR"]
    config.BACKTEST_INI_DIR = paths["BACKTEST_INI_DIR"]
    config.BASE_RESULT_DIR = paths["BASE_RESULT_DIR"]
    config.RAW_REPORT_DIR = paths["RAW_REPORT_DIR"]
    return paths


def cleanup_html_png_files():
    """Xóa tất cả file HTML & PNG được tạo ra từ backtest"""
    cleanup_dirs = [
        MT5_TERMINAL_ROOT,
        MT5_TESTER_FILES,
        MT5_TESTER_REPORTS,
        MT5_FILES,
        MT5_REPORTS,
        BASE_RESULT_DIR,
    ]
    
    removed_count = 0
    for directory in cleanup_dirs:
        if not os.path.isdir(directory):
            continue
        try:
            for file_name in os.listdir(directory):
                if file_name.lower().endswith(('.htm', '.html', '.png')):
                    file_path = os.path.join(directory, file_name)
                    # Bỏ qua một số file đặc biệt
                    if file_name.lower().endswith(('-holding.png', '-hst.png', '-mfemae.png')):
                        continue
                    try:
                        os.remove(file_path)
                        removed_count += 1
                    except Exception as e:
                        print(f"[CLEANUP] Failed to delete {file_name}: {e}", flush=True)
        except Exception as e:
            print(f"[CLEANUP] Error scanning {directory}: {e}", flush=True)
    
    if removed_count > 0:
        print(f"[CLEANUP] Deleted {removed_count} HTML/PNG files", flush=True)


def kill_mt5(pid: int = None):
    """Kill MT5 process và chờ đến khi thực sự exit."""
    if pid:
        config.kill_mt5(pid=pid)
        # Chờ thêm để process fully release file handles
        time.sleep(5)
        
        # Verify process thực sự đã exit
        for attempt in range(10):
            try:
                os.kill(pid, 0)  # signal 0 = check process still alive
                time.sleep(0.5)
            except OSError:
                # Process không tồn tại - OK
                break
    else:
        config.kill_mt5(pid=pid)
        time.sleep(5)


def patch_ini_add_report(ini_path: str, report_name: str) -> str:
    for enc in ["utf-16", "utf-8-sig", "utf-8", "latin-1"]:
        try:
            with open(ini_path, "r", encoding=enc) as f:
                lines = f.readlines()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        raise RuntimeError(f"Khong the doc file ini: {ini_path}")

    has_report = has_replace = has_shutdown = False
    report_file_name = f"{report_name}.htm"
    new_lines = []
    for line in lines:
        s = line.strip().lower()
        if s.startswith("report="):
            new_lines.append(f"Report={report_file_name}\n")
            has_report = True
        elif s.startswith("replacereport="):
            new_lines.append("ReplaceReport=1\n")
            has_replace = True
        elif s.startswith("shutdownterminal="):
            new_lines.append("ShutdownTerminal=1\n")
            has_shutdown = True
        else:
            new_lines.append(line)

    insert_lines = []
    if not has_report:
        insert_lines.append(f"Report={report_file_name}\n")
    if not has_replace:
        insert_lines.append("ReplaceReport=1\n")
    if not has_shutdown:
        insert_lines.append("ShutdownTerminal=1\n")

    if insert_lines:
        idx = len(new_lines)
        in_t = False
        for i, line in enumerate(new_lines):
            if line.strip().lower() == "[tester]":
                in_t = True
            elif in_t and line.strip().startswith("[") and line.strip().lower() != "[tester]":
                idx = i
                break
        for j, insert_line in enumerate(insert_lines):
            new_lines.insert(idx + j, insert_line)

    tmp = ini_path.replace(".ini", "_tmp.ini")
    with open(tmp, "w", encoding="utf-16") as f:
        f.writelines(new_lines)
    return tmp


def find_report_file(report_name: str):
    candidates = [
        f"{report_name}.xml",
        f"{report_name}.xml.htm",
        f"{report_name}.htm",
        report_name,
    ]
    for directory in SEARCH_DIRS:
        if not os.path.isdir(directory):
            continue
        for candidate_name in candidates:
            candidate = os.path.join(directory, candidate_name)
            if os.path.exists(candidate) and os.path.getsize(candidate) > 0:
                ext = os.path.splitext(candidate_name)[1].lower()
                print(f"[FOUND] Report file: {candidate}", flush=True)
                return candidate, ext if ext else ".xml"
    return None, None


def wait_for_report(report_name: str, proc, timeout: int):
    start = time.time()
    check_count = 0
    while (time.time() - start) < timeout:
        check_count += 1
        report_path, report_ext = find_report_file(report_name)
        if report_path:
            print(f"[REPORT] Found: {report_path}", flush=True)
            time.sleep(2)
            return report_path, report_ext
        
        # Check if MT5 process crashed
        exit_code = proc.poll()
        if exit_code is not None:
            print(f"[MT5] Process exited with code {exit_code}. Checking report one more time...", flush=True)
            time.sleep(10)
            return find_report_file(report_name)
        
        # Log every 12 checks (60 seconds)
        if check_count % 12 == 0:
            elapsed = int(time.time() - start)
            print(f"[WAIT] Still waiting for report... ({elapsed}s elapsed, searching in {len(SEARCH_DIRS)} directories)", flush=True)
        
        time.sleep(5)
    
    # Timeout - log where we searched
    print(f"[TIMEOUT] Report not found after {timeout}s for {report_name}", flush=True)
    print(f"[DEBUG] Searched in directories:", flush=True)
    for directory in SEARCH_DIRS:
        exists = os.path.isdir(directory)
        print(f"  - {directory} (exists: {exists})", flush=True)
    return None, None


def extract_bot_info(folder_name: str):
    clean = folder_name.replace("Filtered_", "").replace("BT_Set_", "")
    bot_id = ""
    symbol = ""
    timeframe = ""

    id_match = re.search(r"(\d+)", clean)
    if id_match:
        bot_id = id_match.group(1)

    tf_match = re.search(r"(M15|M30|H1|H4|D1)", clean.upper())
    if tf_match:
        timeframe = tf_match.group(1)

    common_symbols = ["DE30", "HK50", "USOIL", "BTCUSD", "ETHUSD", "XAUUSD", "XAGUSD", "US500", "JP225"]
    upper_clean = clean.upper()
    for s in common_symbols:
        if s in upper_clean:
            symbol = s
            break

    return bot_id, symbol, timeframe


def extract_profit_from_name(report_name: str):
    m = re.search(r"_(\d+_\d+)$", report_name)
    if m:
        return m.group(1)
    m = re.search(r"Profit_(\d+)", report_name)
    return m.group(1) if m else "0"


def parse_report_to_tables(src_path: str):
    from bs4 import BeautifulSoup

    for enc in ["utf-16", "utf-8-sig", "utf-8", "latin-1"]:
        try:
            with open(src_path, "r", encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        raise RuntimeError(f"Khong doc duoc file report: {src_path}")

    soup = BeautifulSoup(content, "lxml")
    tables = soup.find_all("table")
    if not tables:
        raise RuntimeError(f"Khong tim thay bang report trong: {src_path}")

    result = {"summary": [], "orders": [], "deals": []}

    for r in tables[0].find_all("tr"):
        cells = [c.get_text(strip=True) for c in r.find_all(["td", "th"])]
        if any(cells):
            result["summary"].append(cells)

    if len(tables) >= 2:
        current = header = None
        data = []
        sections = {}
        for r in tables[1].find_all("tr"):
            cells = [c.get_text(strip=True) for c in r.find_all(["td", "th"])]
            if not any(cells):
                continue
            ne = [c for c in cells if c]
            if len(ne) == 1 and ne[0] in ("Orders", "Deals", "Positions"):
                if current:
                    sections[current] = (header, data)
                current, header, data = ne[0], None, []
                continue
            if current:
                if header is None:
                    header = cells
                else:
                    data.append(cells)
        if current:
            sections[current] = (header, data)
        for sec, (hdr, rows_data) in sections.items():
            key = sec.lower() if sec.lower() in result else "deals"
            result[key] = ([hdr] if hdr else []) + rows_data

    return result


def to_number(val):
    if not isinstance(val, str) or not val.strip():
        return val
    cleaned = val.replace("\xa0", "").replace(" ", "").replace(",", "")
    try:
        f = float(cleaned)
        return int(f) if f == int(f) else round(f, 8)
    except ValueError:
        return val


def save_xlsx_from_report(report_path: str, report_name: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    data = parse_report_to_tables(report_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092")
    center = Alignment(horizontal="center", vertical="center")
    section_font = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="DCE6F1")

    cur = 1

    for cells in data.get("summary", []):
        ne = [c for c in cells if c]
        if len(ne) == 1 and ne[0] in ("Settings", "Results"):
            cell = ws.cell(row=cur, column=1, value=ne[0])
            cell.font = section_font
            cell.fill = section_fill
        else:
            for c_idx, val in enumerate(cells, 1):
                ws.cell(row=cur, column=c_idx, value=to_number(val))
        cur += 1

    cur += 2
    image_path = None
    for directory in [MT5_TESTER_REPORTS, MT5_TERMINAL_ROOT, MT5_TESTER_FILES, MT5_FILES, MT5_REPORTS, BASE_RESULT_DIR]:
        if not os.path.isdir(directory):
            continue
        candidate = os.path.join(directory, f"{report_name}.png")
        if os.path.exists(candidate) and os.path.getsize(candidate) > 0:
            image_path = candidate
            break

    if image_path:
        try:
            img = XLImage(image_path)
            img.anchor = f"A{cur}"
            ws.add_image(img)
            cur += max(18, int(getattr(img, 'height', 300) / 15)) + 2
        except Exception:
            pass

    if data.get("orders"):
        cell = ws.cell(row=cur, column=1, value="Orders")
        cell.font = section_font
        cell.fill = section_fill
        cur += 1
        for r_idx, row in enumerate(data["orders"]):
            is_hdr = (r_idx == 0)
            for c_idx, val in enumerate(row, 1):
                v = val if is_hdr else to_number(val)
                cell = ws.cell(row=cur, column=c_idx, value=v)
                if is_hdr:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
            cur += 1
        cur += 1

    if data.get("deals"):
        cell = ws.cell(row=cur, column=1, value="Deals")
        cell.font = section_font
        cell.fill = section_fill
        cur += 1
        for r_idx, row in enumerate(data["deals"]):
            is_hdr = (r_idx == 0)
            for c_idx, val in enumerate(row, 1):
                v = val if is_hdr else to_number(val)
                cell = ws.cell(row=cur, column=c_idx, value=v)
                if is_hdr:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
            cur += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    bot_id, symbol, timeframe = extract_bot_info(report_name)
    profit = extract_profit_from_name(report_name)
    out_name = f"{bot_id}{symbol}{timeframe}_{profit}.xlsx"
    out_path = os.path.join(DIR_XLSX, out_name)
    wb.save(out_path)
    return out_path


def cleanup_report_bundle(report_name: str, keep_path: str = None):
    keep_path_normalized = os.path.normcase(os.path.abspath(keep_path)) if keep_path else None
    deleted_count = 0
    failed_count = 0
    
    for directory in [MT5_TESTER_REPORTS, MT5_TERMINAL_ROOT, MT5_TESTER_FILES, MT5_FILES, MT5_REPORTS, BASE_RESULT_DIR]:
        if not os.path.isdir(directory):
            continue
        try:
            for file_name in os.listdir(directory):
                if not file_name.startswith(report_name):
                    continue
                # Xóa tất cả HTML & PNG files liên quan đến report này
                if not file_name.lower().endswith(('.htm', '.html', '.png')):
                    continue
                
                junk = os.path.join(directory, file_name)
                junk_normalized = os.path.normcase(os.path.abspath(junk))
                if keep_path_normalized and junk_normalized == keep_path_normalized:
                    continue
                
                # Retry deleting
                for attempt in range(3):
                    try:
                        if os.path.exists(junk):
                            os.remove(junk)
                            deleted_count += 1
                            break
                    except Exception as e:
                        if attempt == 2:
                            print(f"[CLEANUP] Failed to delete {file_name}: {e}", flush=True)
                            failed_count += 1
                        else:
                            time.sleep(0.5)
        except Exception as e:
            print(f"[CLEANUP] Error scanning directory {directory}: {e}", flush=True)
    
    if deleted_count > 0 or failed_count > 0:
        print(f"[CLEANUP] Cleaned up: {deleted_count} deleted, {failed_count} failed", flush=True)


# ============================================================
# HEARTBEAT: In log mỗi 30 giây để Electron không nghĩ process bị treo
# (Electron sẽ force-kill nếu stdout im lặng >120 giây)
# ============================================================
class HeartbeatThread:
    """In '[HEARTBEAT] still running...' mỗi 30s để giữ stdout sống."""
    def __init__(self, interval: int = 30):
        self._interval = interval
        self._stop_event = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)

    def start(self):
        self._thread.start()

    def stop(self):
        self._stop_event.set()
        self._thread.join(timeout=5)

    def _run(self):
        while not self._stop_event.wait(self._interval):
            print(f"[HEARTBEAT] Backtest dang chay... (moi {self._interval}s)", flush=True)


def run_auto(terminal_data_path=None, mt5_path=None):
    global MT5_PATH, DIR_XLSX, SEARCH_DIRS, MT5_TERMINAL_ROOT, MT5_TESTER_FILES, MT5_TESTER_REPORTS, MT5_FILES, MT5_REPORTS, BASE_RESULT_DIR, INI_BASE_DIR
    
    # Cập nhật MT5_PATH nếu được cung cấp
    if mt5_path:
        MT5_PATH = mt5_path
        print(f"[INIT] Using custom MT5_PATH: {MT5_PATH}", flush=True)
    else:
        print(f"[INIT] Using default MT5_PATH: {MT5_PATH}", flush=True)
    
    build_paths(terminal_data_path)
    
    # ⚠️ CRITICAL: Cập nhật lại SEARCH_DIRS sau build_paths() để match terminal_data_path của profile
    MT5_TERMINAL_ROOT = config.TERMINAL_DATA_PATH
    MT5_TESTER_FILES = os.path.join(MT5_TERMINAL_ROOT, r"Tester\Files")
    MT5_TESTER_REPORTS = os.path.join(MT5_TERMINAL_ROOT, r"Tester\Reports")
    MT5_FILES = os.path.join(MT5_TERMINAL_ROOT, r"MQL5\Files")
    MT5_REPORTS = os.path.join(MT5_TERMINAL_ROOT, r"reports")
    BASE_RESULT_DIR = config.BASE_RESULT_DIR
    INI_BASE_DIR = config.BACKTEST_INI_DIR
    SEARCH_DIRS = [MT5_TESTER_REPORTS, MT5_TESTER_FILES, MT5_FILES, MT5_REPORTS, BASE_RESULT_DIR, MT5_TERMINAL_ROOT]
    
    # Cập nhật DIR_XLSX và INI_BASE_DIR dựa trên config mới (sau build_paths)
    DIR_XLSX = os.path.join(config.BASE_RESULT_DIR, "xlsx")
    INI_BASE_DIR = config.BACKTEST_INI_DIR
    os.makedirs(DIR_XLSX, exist_ok=True)
    print(f"[INIT] Output dir: {DIR_XLSX}", flush=True)
    print(f"[INIT] INI source dir: {INI_BASE_DIR}", flush=True)
    
    # Cleanup HTML & PNG files cũ từ lần chạy trước
    print("[INIT] Cleaning up old HTML/PNG files...", flush=True)
    cleanup_html_png_files()
    
    all_inis = []
    for root, dirs, files in os.walk(INI_BASE_DIR):
        for file_name in files:
            if file_name.endswith(".ini") and not file_name.endswith("_tmp.ini"):
                all_inis.append(os.path.join(root, file_name))

    print(f"[INIT] Tong cong {len(all_inis)} file. Bat dau chay TUAN TU...\n", flush=True)
    
    # Log first INI file structure for debugging
    if all_inis:
        print(f"[DEBUG] First INI: {all_inis[0]}", flush=True)
        try:
            with open(all_inis[0], 'r', encoding='utf-16') as f:
                head = [f.readline() for _ in range(5)]
                print(f"[DEBUG] First 5 lines: {head}", flush=True)
        except Exception as e:
            print(f"[DEBUG] Error reading INI: {e}", flush=True)
    
    if len(all_inis) == 0:
        print("\n" + "=" * 55)
        print("❌ ERROR: Khong tim thay file INI nao!")
        print("=" * 55)
        print(f"📁 Searching in: {INI_BASE_DIR}")
        print(f"📁 Directory exists: {os.path.isdir(INI_BASE_DIR)}")
        if os.path.isdir(INI_BASE_DIR):
            try:
                items = os.listdir(INI_BASE_DIR)
                print(f"📁 Items found: {len(items)}")
                if items:
                    print(f"   First 10 items: {items[:10]}")
            except Exception as e:
                print(f"📁 Error listing directory: {e}")
        print("\n💡 Solutions:")
        print("  1. Chạy OPTIMIZE trước để tạo INI files")
        print("  2. Hoặc sao chép INI files vào folder trên")
        print("=" * 55 + "\n")
        sys.exit(1)

    success_list = []
    fail_list    = []

    # Khởi động heartbeat để Electron không force-kill do stdout im lặng >120s
    heartbeat = HeartbeatThread(interval=30)
    heartbeat.start()

    try:
        # ============================================================
        # PHASE 1-3: Chạy TUẦN TỰ từng INI trong một terminal
        # ============================================================
        for idx, path_ini in enumerate(all_inis, 1):
            report_name = os.path.basename(path_ini).replace('.ini', '')
            print(f"[PHASE1] [{idx}/{len(all_inis)}] Prepare backtest for: {report_name}", flush=True)

            tmp_ini = patch_ini_add_report(path_ini, report_name)
            print(f"[DEBUG] Temp INI created: {tmp_ini} (exists: {os.path.exists(tmp_ini)})", flush=True)
            
            proc = None
            mt5_pid = None
            try:
                print(f"[PHASE1] [{idx}/{len(all_inis)}] Spawn MT5", flush=True)
                print(f"[DEBUG] MT5 path: {MT5_PATH}", flush=True)
                print(f"[DEBUG] Config INI: {tmp_ini}", flush=True)
                # Use shell=False with list args - same as optimize script
                proc = subprocess.Popen(
                    [MT5_PATH, f'/config:{tmp_ini}'],
                    shell=False
                )
                mt5_pid = proc.pid
                print(f"[PHASE1] [{idx}/{len(all_inis)}] MT5 spawned - PID={mt5_pid}", flush=True)

                print(f"[PHASE2] [{idx}/{len(all_inis)}] Doi report cho: {report_name}", flush=True)
                found_path, _ = wait_for_report(report_name, proc, TIMEOUT_SECONDS)

                if not found_path:
                    print(f"[PHASE2] [{idx}/{len(all_inis)}] Timeout ({TIMEOUT_SECONDS}s): {report_name}", flush=True)
                    fail_list.append(report_name)
                    continue

                print(f"[PHASE2] [{idx}/{len(all_inis)}] Found report: {report_name}", flush=True)

                print(f"[PHASE3] [{idx}/{len(all_inis)}] Dang luu XLSX...", flush=True)
                xlsx_path = save_xlsx_from_report(found_path, report_name)
                print(f"[PHASE3] [{idx}/{len(all_inis)}] Saved Excel: {xlsx_path}", flush=True)
                success_list.append(report_name)
                cleanup_report_bundle(report_name, keep_path=found_path)

            except Exception as e:
                print(f"[PHASE3] [{idx}/{len(all_inis)}] Error: {e}", flush=True)
                fail_list.append(report_name)
            finally:
                if proc is not None:
                    try:
                        kill_mt5(pid=mt5_pid)
                        print(f"[PHASE4] [{idx}/{len(all_inis)}] Killed MT5 PID={mt5_pid} for {report_name}", flush=True)
                    except Exception as e:
                        print(f"[PHASE4] [{idx}/{len(all_inis)}] Failed to kill {report_name}: {e}", flush=True)
                
                # Delay để đảm bảo MT5 process fully released file lock
                time.sleep(2)
                
                # Xóa tmp_ini với retry logic
                for attempt in range(3):
                    try:
                        if os.path.exists(tmp_ini):
                            os.remove(tmp_ini)
                            print(f"[CLEANUP] [{idx}/{len(all_inis)}] Deleted tmp ini", flush=True)
                            break
                    except Exception as e:
                        print(f"[CLEANUP] [{idx}/{len(all_inis)}] Failed to delete tmp ini (attempt {attempt+1}/3): {e}", flush=True)
                        if attempt < 2:
                            time.sleep(1)

        # ============================================================
        # PHASE 5: Done
        # ============================================================
        print(f"\n[PHASE5] Done processing all reports.\n", flush=True)

    finally:
        # Dừng heartbeat dù thành công hay lỗi
        heartbeat.stop()

    # Cleanup HTML & PNG files tạo ra từ backtest
    print("\n[CLEANUP] Cleaning up generated HTML/PNG files...", flush=True)
    cleanup_html_png_files()

    print("\n" + "=" * 55)
    print(f"Thanh cong : {len(success_list)}/{len(all_inis)}")
    print(f"That bai   : {len(fail_list)}/{len(all_inis)}")
    print("=" * 55)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='SaveBacktest Optimizer')
    parser.add_argument('--mt5-path', type=str, default=None, help='Duong dan MT5 Executable (tuy chon)')
    parser.add_argument('--terminal-path', type=str, default=None, help='Duong dan Terminal Data (tuy chon)')
    args = parser.parse_args()
    run_auto(args.terminal_path, args.mt5_path)