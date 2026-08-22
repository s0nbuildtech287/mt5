import os
import sys
import re
import time
from glob import glob
from datetime import datetime

# Tự động kiểm tra và cài đặt thư viện cần thiết nếu chưa có
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
except ImportError:
    print("Đang cài đặt thư viện cần thiết (google-api-python-client, google-auth, openpyxl)...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "google-api-python-client", "google-auth", "openpyxl"])
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

SERVICE_ACCOUNT_DIR = r"C:\Users\SonBx\Desktop\Lotusquant\Optimize\backend\src\config\serviceAccounts"
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

VALID_SYMBOLS = [
    "BTCUSD", "ETHUSD", "XAUUSD", "XAGUSD", "US500", "DE30", "JP225", "HK50",
    "USOIL", "STOXX50", "UK100", "US30", "USTECH", "NAS100", "UKOIL", "DE40",
    "EURUSD", "GBPUSD", "AUDUSD", "NZDUSD", "USDCAD", "USDCHF", "USDJPY"
]
VALID_TIMEFRAMES = ["H1", "H4"]
VALID_ORDER_TYPES = ["OB", "OS"]

SEP = "=" * 105


# =============================================================================
# DRIVE HELPERS
# =============================================================================

class DriveAccountManager:
    """Quản lý và xoay vòng nhiều Service Account để tránh dính quota limit"""
    def __init__(self, sa_dir):
        self.sa_files = glob(os.path.join(sa_dir, "*.json"))
        if not self.sa_files:
            raise FileNotFoundError(f"Không tìm thấy file service account nào trong thư mục: {sa_dir}")
        self.current_idx = 0
        print(f"Đã tìm thấy {len(self.sa_files)} Service Accounts trong {sa_dir}")

    def get_service(self):
        sa_file = self.sa_files[self.current_idx]
        creds = service_account.Credentials.from_service_account_file(sa_file, scopes=SCOPES)
        return build('drive', 'v3', credentials=creds)

    def rotate_account(self):
        self.current_idx = (self.current_idx + 1) % len(self.sa_files)
        return self.get_service()


def extract_folder_id(url_or_id):
    """Trích xuất Folder ID từ đường dẫn Google Drive hoặc ID gốc"""
    url_or_id = url_or_id.strip()
    match = re.search(r'folders/([a-zA-Z0-9_-]+)', url_or_id)
    if match:
        return match.group(1)
    match_id = re.search(r'id=([a-zA-Z0-9_-]+)', url_or_id)
    if match_id:
        return match_id.group(1)
    return url_or_id


def list_items_in_folder(service_mgr, folder_id):
    """Lấy tất cả items (subfolders và files) trong 1 folder"""
    items = []
    page_token = None
    query = f"'{folder_id}' in parents and trashed = false"
    service = service_mgr.get_service()

    while True:
        try:
            response = service.files().list(
                q=query,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields='nextPageToken, files(id, name, mimeType, size, modifiedTime)',
                pageToken=page_token,
                pageSize=1000
            ).execute()
            items.extend(response.get('files', []))
            page_token = response.get('nextPageToken')
            if not page_token:
                break
        except HttpError as error:
            if error.resp.status in [403, 429]:
                print(f"\n[QUOTA EXCEEDED] Đang xoay vòng Service Account...")
                service = service_mgr.rotate_account()
                time.sleep(1)
            else:
                print(f"\nLỗi API Google Drive: {error}")
                break
        except Exception as e:
            print(f"\nLỗi không xác định: {e}")
            break

    return items


def scan_asset_folder(service_mgr, folder_id, current_path=""):
    """Duyệt đệ quy các item trong folder tài sản, phát hiện file/folder lạ"""
    all_files = []
    unexpected_items = []
    items = list_items_in_folder(service_mgr, folder_id)

    EXCEL_MIME_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/vnd.google-apps.spreadsheet"
    }

    for item in items:
        mime_type = item.get('mimeType', '')
        item_name = item.get('name', '')
        item_id = item.get('id', '')
        item_path = f"{current_path}/{item_name}" if current_path else item_name

        if mime_type == 'application/vnd.google-apps.folder':
            unexpected_items.append({
                "id": item_id,
                "name": item_name,
                "type": "Folder lạ trong Folder tài sản",
                "reason": "Phát hiện thư mục con lạ nằm bên trong folder tài sản",
                "path": item_path
            })
            sub_files, sub_unexpected = scan_asset_folder(service_mgr, item_id, item_path)
            all_files.extend(sub_files)
            unexpected_items.extend(sub_unexpected)
        else:
            item['path'] = item_path
            all_files.append(item)
            is_excel = item_name.lower().endswith(('.xlsx', '.xls')) or mime_type in EXCEL_MIME_TYPES
            if not is_excel:
                unexpected_items.append({
                    "id": item_id,
                    "name": item_name,
                    "type": "File lạ (Không phải file Excel)",
                    "reason": f"File không phải định dạng Excel/Google Sheet (MIME: {mime_type})",
                    "path": item_path
                })

    return all_files, unexpected_items


# =============================================================================
# SCAN DRIVE — dùng chung cho cả 2 option
# =============================================================================

def scan_drive(url_or_id, service_mgr):
    """
    Quét Drive, trả về:
      all_files        - toàn bộ file tìm thấy (đã có bot_id, symbol, path)
      asset_folders    - danh sách folder tài sản
      report_data      - thống kê theo folder (chưa có valid/invalid name)
      bot_map          - {bot_id: [files]}
      raw_unexpected   - file/folder lạ phát hiện ở bước đếm (folder không chuẩn, file không excel)
    """
    folder_id = extract_folder_id(url_or_id)
    print(f"\n  Drive Folder ID: {folder_id}")
    print("  Đang quét cấu trúc thư mục và file trên Drive...")

    top_items = list_items_in_folder(service_mgr, folder_id)
    root_files = [i for i in top_items if i.get('mimeType') != 'application/vnd.google-apps.folder']
    asset_folders = [i for i in top_items if i.get('mimeType') == 'application/vnd.google-apps.folder']

    all_files = []
    raw_unexpected = []
    report_data = []
    bot_map = {}

    EXCEL_MIME_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/vnd.google-apps.spreadsheet"
    }

    # File nằm thẳng ở root Drive (lạ)
    for rf in root_files:
        rf_name = rf.get('name', '')
        raw_unexpected.append({
            "id": rf.get('id'),
            "name": rf_name,
            "type": "File lạ tại Drive Root",
            "reason": "File nằm trực tiếp ở thư mục gốc Drive (không nằm trong folder tài sản nào)",
            "path": f"/[Drive Root]/{rf_name}"
        })
        all_files.append(rf)

    for idx, asset in enumerate(asset_folders, 1):
        asset_id = asset.get('id')
        asset_name = asset.get('name', '').strip()
        asset_path = asset_name

        # Folder tài sản có tên không chứa Symbol chuẩn
        asset_upper = asset_name.upper()
        if not any(sym in asset_upper for sym in VALID_SYMBOLS):
            raw_unexpected.append({
                "id": asset_id,
                "name": asset_name,
                "type": "Folder lạ tại Drive Root",
                "reason": "Tên folder tài sản không chứa mã Symbol chuẩn (Ví dụ: BTCUSD, XAUUSD,...)",
                "path": f"/[Drive Root]/{asset_name}"
            })

        print(f"  [{idx}/{len(asset_folders)}] Đang quét folder: {asset_name} ...", end="", flush=True)
        files_in_asset, sub_unexpected = scan_asset_folder(service_mgr, asset_id, asset_path)
        raw_unexpected.extend(sub_unexpected)

        xlsx_count = sum(
            1 for f in files_in_asset
            if f.get('name', '').lower().endswith(('.xlsx', '.xls')) or f.get('mimeType') in EXCEL_MIME_TYPES
        )

        for f in files_in_asset:
            fname = f.get('name', '')
            bot_id_match = re.match(r'^(\d{3})', fname)
            bot_id = bot_id_match.group(1) if bot_id_match else None
            f['bot_id'] = bot_id or "-"

            extracted_sym = "-"
            fname_upper = fname.upper()
            for sym in VALID_SYMBOLS:
                if sym in fname_upper:
                    extracted_sym = sym
                    break
            f['symbol'] = extracted_sym

            if bot_id:
                bot_map.setdefault(bot_id, []).append(f)

        all_files.extend(files_in_asset)
        report_data.append({
            "folder": asset_name,
            "total_files": len(files_in_asset),
            "xlsx_files": xlsx_count,
            "raw_files_in_asset": files_in_asset
        })
        print(f" -> Done ({len(files_in_asset)} files)")

    return all_files, asset_folders, report_data, bot_map, raw_unexpected


def build_bot_summary(bot_map):
    bot_summary = []
    for b_id in sorted(bot_map.keys()):
        b_files = bot_map[b_id]
        symbols = sorted(set(f['symbol'] for f in b_files if f['symbol'] != "-"))
        bot_summary.append({
            "bot_id": b_id,
            "file_count": len(b_files),
            "symbols": symbols if symbols else ["Không xác định"],
            "status": "Đủ chiến lược" if len(b_files) >= 4 else "Ít chiến lược"
        })
    return bot_summary


# =============================================================================
# CHECK FILENAME FORMAT
# =============================================================================

def check_filename_format(name):
    """
    Kiểm tra tên file đúng chuẩn: {BotId}{Symbol}{Timeframe}{OrderType}
    Ví dụ hợp lệ: 001BTCUSDH1OB.xlsx
    """
    base = re.sub(r'\.[^/.]+$', '', name)

    if base != base.upper():
        return False, "Tên file phải viết HOA toàn bộ", None, None

    bot_id_match = re.match(r'^(\d{3})', base)
    if not bot_id_match:
        return False, "BotId không hợp lệ (phải là 3 chữ số đầu tiên)", None, None

    bot_id = bot_id_match.group(1)
    rest = base[3:]

    symbol = None
    after_symbol = rest
    for sym in VALID_SYMBOLS:
        if rest.startswith(sym):
            symbol = sym
            after_symbol = rest[len(sym):]
            break

    if not symbol:
        return False, "Symbol không nằm trong danh sách hợp lệ", bot_id, None

    timeframe = None
    after_tf = after_symbol
    for tf in VALID_TIMEFRAMES:
        if after_symbol.startswith(tf):
            timeframe = tf
            after_tf = after_symbol[len(tf):]
            break

    if not timeframe:
        return False, "Khung thời gian không hợp lệ (phải là H1 hoặc H4)", bot_id, symbol

    order_type = None
    for ot in VALID_ORDER_TYPES:
        if after_tf == ot:
            order_type = ot
            break

    if not order_type:
        return False, "Kiểu lệnh không hợp lệ (phải là OB hoặc OS)", bot_id, symbol

    return True, "Hợp lệ", bot_id, symbol


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def _apply_header_style(ws):
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center


def _autofit_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def export_option1_excel(output_filename, report_data, bot_summary, unexpected_items, all_files):
    """Xuất Excel cho Option 1: Thống kê đếm + file/folder lạ"""
    wb = openpyxl.Workbook()
    err_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    err_font = Font(name="Calibri", size=11, color="C00000", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Thống kê theo folder tài sản
    ws1 = wb.active
    ws1.title = "Thong_Ke_Tai_San"
    ws1.append(["STT", "Tên Folder Tài Sản", "Tổng Số File", "Số File .xlsx", "Số Mục Lạ"])
    _apply_header_style(ws1)
    for idx, row in enumerate(report_data, 1):
        ws1.append([
            idx,
            row['folder'],
            row['total_files'],
            row['xlsx_files'],
            row.get('unexpected_count', 0)
        ])
        ws1.cell(row=idx + 1, column=1).alignment = align_center
        ws1.cell(row=idx + 1, column=3).alignment = align_center
        ws1.cell(row=idx + 1, column=4).alignment = align_center
        ws1.cell(row=idx + 1, column=5).alignment = align_center
        if row.get('unexpected_count', 0) > 0:
            ws1.cell(row=idx + 1, column=5).font = err_font
            ws1.cell(row=idx + 1, column=5).fill = err_fill

    # Sheet 2: Thống kê Bot
    ws_bot = wb.create_sheet(title="Thong_Ke_Danh_Sach_Bot")
    ws_bot.append(["STT", "Mã Con Bot (Bot ID)", "Số Lượng Chiến Lược (File)", "Danh Sách Symbol", "Trạng Thái"])
    _apply_header_style(ws_bot)
    for idx, bot in enumerate(bot_summary, 1):
        ws_bot.append([idx, f"Bot {bot['bot_id']}", bot['file_count'], ", ".join(bot['symbols']), bot['status']])
        for col in [1, 2, 3, 5]:
            ws_bot.cell(row=idx + 1, column=col).alignment = align_center

    # Sheet 3: Danh sách mục lạ
    ws_err = wb.create_sheet(title="File_Folder_La")
    ws_err.append(["STT", "Tên Mục / File", "Loại", "Mô Tả / Lý Do", "Đường Dẫn Trong Drive", "File ID"])
    _apply_header_style(ws_err)
    if not unexpected_items:
        ws_err.append([1, "-", "OK", "Không phát hiện file/folder lạ nào!", "-", "-"])
    else:
        for idx, item in enumerate(unexpected_items, 1):
            ws_err.append([
                idx,
                item.get('name', ''),
                item.get('type', ''),
                item.get('reason', ''),
                item.get('path', ''),
                item.get('id', '')
            ])
            ws_err.cell(row=idx + 1, column=1).alignment = align_center
            ws_err.cell(row=idx + 1, column=2).font = err_font
            ws_err.cell(row=idx + 1, column=2).fill = err_fill

    # Sheet 4: Toàn bộ file
    ws_all = wb.create_sheet(title="Danh_Sach_Toan_Bo_File")
    ws_all.append(["STT", "Tên File", "Mã Bot", "Symbol", "Đường Dẫn Trong Drive", "File ID"])
    _apply_header_style(ws_all)
    for idx, f in enumerate(all_files, 1):
        ws_all.append([idx, f.get('name', ''), f.get('bot_id', '-'), f.get('symbol', '-'), f.get('path', ''), f.get('id', '')])
        for col in [1, 3, 4]:
            ws_all.cell(row=idx + 1, column=col).alignment = align_center

    for ws in [ws1, ws_bot, ws_err, ws_all]:
        _autofit_columns(ws)

    wb.save(output_filename)
    print(f"\n  📂 Đã xuất Excel: {os.path.abspath(output_filename)}")


def export_option2_excel(output_filename, report_data, bot_summary, name_errors, all_files):
    """Xuất Excel cho Option 2: Kết quả check định dạng tên file"""
    wb = openpyxl.Workbook()
    err_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    err_font = Font(name="Calibri", size=11, color="C00000", bold=True)
    ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ok_font = Font(name="Calibri", size=11, color="375623", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Tổng hợp theo folder
    ws1 = wb.active
    ws1.title = "Tong_Hop_Dinh_Dang"
    ws1.append(["STT", "Tên Folder Tài Sản", "Tổng File", "Tên Hợp Lệ", "Tên SAI Định Dạng"])
    _apply_header_style(ws1)
    for idx, row in enumerate(report_data, 1):
        ws1.append([
            idx,
            row['folder'],
            row['total_files'],
            row.get('valid_name_files', 0),
            row.get('invalid_name_files', 0)
        ])
        ws1.cell(row=idx + 1, column=1).alignment = align_center
        ws1.cell(row=idx + 1, column=3).alignment = align_center
        ws1.cell(row=idx + 1, column=4).alignment = align_center
        ws1.cell(row=idx + 1, column=5).alignment = align_center
        if row.get('invalid_name_files', 0) > 0:
            ws1.cell(row=idx + 1, column=5).font = err_font
            ws1.cell(row=idx + 1, column=5).fill = err_fill
        else:
            ws1.cell(row=idx + 1, column=4).font = ok_font
            ws1.cell(row=idx + 1, column=4).fill = ok_fill

    # Sheet 2: Danh sách file SAI định dạng
    ws_err = wb.create_sheet(title="File_Sai_Dinh_Dang")
    ws_err.append(["STT", "Tên File", "Lý Do Sai", "Đường Dẫn", "File ID"])
    _apply_header_style(ws_err)
    if not name_errors:
        ws_err.append([1, "-", "TẤT CẢ TÊN FILE ĐỀU HỢP LỆ!", "-", "-"])
    else:
        for idx, item in enumerate(name_errors, 1):
            ws_err.append([
                idx,
                item.get('name', ''),
                item.get('reason', ''),
                item.get('path', ''),
                item.get('id', '')
            ])
            ws_err.cell(row=idx + 1, column=1).alignment = align_center
            ws_err.cell(row=idx + 1, column=2).font = err_font
            ws_err.cell(row=idx + 1, column=2).fill = err_fill

    # Sheet 3: Toàn bộ file với kết quả check tên
    ws_all = wb.create_sheet(title="Danh_Sach_Toan_Bo_File")
    ws_all.append(["STT", "Tên File", "Mã Bot", "Symbol", "Trạng Thái Tên", "Chi Tiết", "Đường Dẫn", "File ID"])
    _apply_header_style(ws_all)
    for idx, f in enumerate(all_files, 1):
        is_valid = f.get('is_valid_name', True)
        status_text = "✓ Hợp lệ" if is_valid else "✗ SAI định dạng"
        ws_all.append([
            idx,
            f.get('name', ''),
            f.get('bot_id', '-'),
            f.get('symbol', '-'),
            status_text,
            f.get('name_reason', '-'),
            f.get('path', ''),
            f.get('id', '')
        ])
        row_idx = idx + 1
        for col in [1, 3, 4, 5]:
            ws_all.cell(row=row_idx, column=col).alignment = align_center
        if not is_valid:
            ws_all.cell(row=row_idx, column=5).font = err_font
            ws_all.cell(row=row_idx, column=5).fill = err_fill
        else:
            ws_all.cell(row=row_idx, column=5).font = ok_font
            ws_all.cell(row=row_idx, column=5).fill = ok_fill

    # Sheet 4: Bot summary
    ws_bot = wb.create_sheet(title="Thong_Ke_Bot")
    ws_bot.append(["STT", "Mã Con Bot (Bot ID)", "Số Lượng File", "Danh Sách Symbol", "Trạng Thái"])
    _apply_header_style(ws_bot)
    for idx, bot in enumerate(bot_summary, 1):
        ws_bot.append([idx, f"Bot {bot['bot_id']}", bot['file_count'], ", ".join(bot['symbols']), bot['status']])
        for col in [1, 2, 3, 5]:
            ws_bot.cell(row=idx + 1, column=col).alignment = align_center

    for ws in [ws1, ws_err, ws_all, ws_bot]:
        _autofit_columns(ws)

    wb.save(output_filename)
    print(f"\n  📂 Đã xuất Excel: {os.path.abspath(output_filename)}")


# =============================================================================
# OPTION 1: Đếm file + phát hiện file/folder lạ
# =============================================================================

def run_option1(drive_links, service_mgr):
    print(f"\n{SEP}")
    print("  OPTION 1 — ĐẾM SỐ LƯỢNG FILE & PHÁT HIỆN FILE / FOLDER LẠ")
    print(SEP)

    for link in drive_links:
        print(f"\n🔗 Đang xử lý: {link.strip()}")
        try:
            all_files, asset_folders, report_data, bot_map, raw_unexpected = scan_drive(link, service_mgr)
        except Exception as e:
            print(f"  ❌ Lỗi khi quét Drive: {e}")
            continue

        bot_summary = build_bot_summary(bot_map)
        total_xlsx = sum(r['xlsx_files'] for r in report_data)

        # Đếm mục lạ theo từng folder
        for row in report_data:
            asset_prefix = row['folder']
            row['unexpected_count'] = sum(
                1 for item in raw_unexpected
                if item.get('path', '').startswith(asset_prefix)
            )

        # --- In kết quả ---
        print(f"\n{SEP}")
        print(f"  📊 TỔNG QUAN")
        print(SEP)
        print(f"  Tổng số file    : {len(all_files)}")
        print(f"  Số file .xlsx   : {total_xlsx}")
        print(f"  Số folder tài sản: {len(asset_folders)}")
        print(f"  Số bot duy nhất : {len(bot_summary)}")
        print(f"  Mục lạ phát hiện: {len(raw_unexpected)}")

        print(f"\n{SEP}")
        print("  📁 THỐNG KÊ THEO FOLDER TÀI SẢN")
        print(SEP)
        print(f"  {'STT':<5} {'Folder':<30} {'Tổng':<8} {'.xlsx':<8} {'Mục lạ':<10}")
        print("  " + "-" * 65)
        for idx, row in enumerate(report_data, 1):
            lạ = row.get('unexpected_count', 0)
            flag = " ⚠️" if lạ > 0 else ""
            print(f"  {idx:<5} {row['folder']:<30} {row['total_files']:<8} {row['xlsx_files']:<8} {lạ}{flag}")

        print(f"\n{SEP}")
        print(f"  🤖 THỐNG KÊ BOT ({len(bot_summary)} bot)")
        print(SEP)
        print(f"  {'Bot ID':<10} {'Số File':<10} {'Trạng Thái':<20} Symbols")
        print("  " + "-" * 70)
        for bot in bot_summary:
            print(f"  {bot['bot_id']:<10} {bot['file_count']:<10} {bot['status']:<20} {', '.join(bot['symbols'])}")

        if raw_unexpected:
            print(f"\n{SEP}")
            print(f"  ⚠️  DANH SÁCH MỤC LẠ / FILE KHÔNG CHUẨN ({len(raw_unexpected)} mục)")
            print(SEP)
            for idx, item in enumerate(raw_unexpected[:30], 1):
                print(f"  {idx:>3}. [{item['type']}] {item['name']}")
                print(f"       → {item['reason']}")
                print(f"       📍 {item['path']}")
            if len(raw_unexpected) > 30:
                print(f"\n  ... và còn {len(raw_unexpected) - 30} mục khác (xem trong file Excel).")
        else:
            print(f"\n  ✅ Không phát hiện file/folder lạ nào!")

        # --- Hỏi xuất Excel ---
        print()
        choice = input("👉 Xuất báo cáo ra file Excel? (y/n, mặc định y): ").strip().lower()
        if choice in ("", "y", "yes"):
            script_dir = os.path.dirname(os.path.abspath(__file__))
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(script_dir, f"Option1_Dem_File_Folder_La_{ts}.xlsx")
            export_option1_excel(out, report_data, bot_summary, raw_unexpected, all_files)
        else:
            print("  ⏩ Bỏ qua xuất Excel.")


# =============================================================================
# OPTION 2: Kiểm tra định dạng tên file
# =============================================================================

def run_option2(drive_links, service_mgr):
    print(f"\n{SEP}")
    print("  OPTION 2 — KIỂM TRA ĐỊNH DẠNG TÊN FILE")
    print(f"  Chuẩn: {{3 số Bot ID}}{{Symbol}}{{H1|H4}}{{OB|OS}}.xlsx   Ví dụ: 001BTCUSDH1OB.xlsx")
    print(SEP)

    for link in drive_links:
        print(f"\n🔗 Đang xử lý: {link.strip()}")
        try:
            all_files, asset_folders, report_data, bot_map, raw_unexpected = scan_drive(link, service_mgr)
        except Exception as e:
            print(f"  ❌ Lỗi khi quét Drive: {e}")
            continue

        bot_summary = build_bot_summary(bot_map)
        name_errors = []

        # Check từng file
        for row in report_data:
            valid_count = 0
            invalid_count = 0
            for f in row['raw_files_in_asset']:
                fname = f.get('name', '')
                is_valid, reason, b_id, sym = check_filename_format(fname)
                f['is_valid_name'] = is_valid
                f['name_reason'] = reason
                if is_valid:
                    valid_count += 1
                else:
                    invalid_count += 1
                    name_errors.append({
                        "id": f.get('id'),
                        "name": fname,
                        "reason": reason,
                        "path": f.get('path', '')
                    })
            row['valid_name_files'] = valid_count
            row['invalid_name_files'] = invalid_count

        total_files = len(all_files)
        total_valid = sum(r['valid_name_files'] for r in report_data)
        total_invalid = len(name_errors)

        # --- In kết quả ---
        print(f"\n{SEP}")
        print("  📊 KẾT QUẢ KIỂM TRA ĐỊNH DẠNG TÊN FILE")
        print(SEP)
        print(f"  Tổng file đã kiểm tra : {total_files}")
        print(f"  ✅ Hợp lệ             : {total_valid}")
        print(f"  ❌ Sai định dạng      : {total_invalid}")

        print(f"\n{SEP}")
        print("  📁 CHI TIẾT THEO FOLDER")
        print(SEP)
        print(f"  {'STT':<5} {'Folder':<30} {'Tổng':<8} {'Hợp lệ':<10} {'Sai':<8}")
        print("  " + "-" * 65)
        for idx, row in enumerate(report_data, 1):
            sai = row.get('invalid_name_files', 0)
            flag = " ⚠️" if sai > 0 else " ✅"
            print(f"  {idx:<5} {row['folder']:<30} {row['total_files']:<8} {row.get('valid_name_files', 0):<10} {sai}{flag}")

        if name_errors:
            print(f"\n{SEP}")
            print(f"  ❌ DANH SÁCH FILE SAI ĐỊNH DẠNG ({total_invalid} file)")
            print(SEP)
            for idx, item in enumerate(name_errors[:40], 1):
                print(f"  {idx:>3}. {item['name']}")
                print(f"       → {item['reason']}")
                print(f"       📍 {item['path']}")
            if total_invalid > 40:
                print(f"\n  ... và còn {total_invalid - 40} file khác (xem trong file Excel).")
        else:
            print(f"\n  🎉 TOÀN BỘ {total_files} FILE ĐỀU CÓ TÊN HỢP LỆ!")

        # --- Hỏi xuất Excel ---
        print()
        choice = input("👉 Xuất báo cáo ra file Excel? (y/n, mặc định y): ").strip().lower()
        if choice in ("", "y", "yes"):
            script_dir = os.path.dirname(os.path.abspath(__file__))
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(script_dir, f"Option2_Check_Dinh_Dang_{ts}.xlsx")
            export_option2_excel(out, report_data, bot_summary, name_errors, all_files)
        else:
            print("  ⏩ Bỏ qua xuất Excel.")


# =============================================================================
# OPTION 3: Lấy tên strategy từ folder tài sản
# =============================================================================

BLOCK_SIZE = 50  # số strategy mỗi block hiển thị


def run_option3(drive_links, service_mgr):
    print(f"\n{SEP}")
    print("  OPTION 3 — LẤY DANH SÁCH TÊN STRATEGY TRONG FOLDER")
    print(f"  Mỗi block hiển thị {BLOCK_SIZE} strategy, ngăn cách bằng dấu phẩy.")
    print(SEP)

    for link in drive_links:
        print(f"\n🔗 Đang xử lý: {link.strip()}")
        folder_id = extract_folder_id(link)
        print(f"  Folder ID: {folder_id}")
        print("  Đang lấy danh sách file...", flush=True)

        try:
            items = list_items_in_folder(service_mgr, folder_id)
        except Exception as e:
            print(f"  ❌ Lỗi khi lấy danh sách: {e}")
            continue

        # Lọc chỉ lấy file (không lấy subfolder), bỏ phần mở rộng
        names = []
        for item in items:
            if item.get('mimeType') == 'application/vnd.google-apps.folder':
                continue
            fname = item.get('name', '')
            # Bỏ phần đuôi file (.xlsx, .xls, ...)
            base = re.sub(r'\.[^/.]+$', '', fname).strip()
            if base:
                names.append(base)

        names.sort()
        total = len(names)

        if total == 0:
            print("  ⚠️  Không tìm thấy file nào trong folder này.")
            continue

        print(f"\n{SEP}")
        print(f"  📋 DANH SÁCH STRATEGY — Tổng cộng: {total} strategy")
        print(SEP)

        # In theo block, mỗi block BLOCK_SIZE tên
        for block_start in range(0, total, BLOCK_SIZE):
            block = names[block_start:block_start + BLOCK_SIZE]
            block_num = block_start // BLOCK_SIZE + 1
            total_blocks = (total + BLOCK_SIZE - 1) // BLOCK_SIZE
            print(f"\n  [Block {block_num}/{total_blocks}  —  #{block_start + 1} đến #{min(block_start + BLOCK_SIZE, total)}]")
            print("  " + ",".join(block))
            print("  " + "-" * 100)

        print(f"\n  Tổng: {total} strategy")


# =============================================================================
# MAIN MENU
# =============================================================================

def print_menu():
    print(f"\n{SEP}")
    print("  VALIDATION TOOL — GOOGLE DRIVE")
    print(SEP)
    print("  1. Đếm số lượng file & phát hiện file/folder lạ")
    print("  2. Kiểm tra định dạng tên file (BotId + Symbol + Timeframe + OrderType)")
    print("  3. Lấy danh sách tên strategy trong folder tài sản")
    print("  0. Thoát")
    print(SEP)


def get_drive_links():
    user_input = input("\nNhập Link hoặc Folder ID Google Drive (nhiều link cách nhau bằng dấu phẩy):\n> ")
    return [link.strip() for link in user_input.split(",") if link.strip()]


def main():
    try:
        service_mgr = DriveAccountManager(SERVICE_ACCOUNT_DIR)
    except Exception as e:
        print(f"\n❌ Lỗi khởi tạo Service Account: {e}")
        return

    while True:
        print_menu()
        choice = input("  Chọn option (0/1/2/3): ").strip()

        if choice == "0":
            print("\n  👋 Thoát. Hẹn gặp lại!\n")
            break

        elif choice in ("1", "2", "3"):
            drive_links = get_drive_links()
            if not drive_links:
                print("\n  ⚠️  Không có link Drive nào được nhập. Thử lại.")
                continue

            if choice == "1":
                run_option1(drive_links, service_mgr)
            elif choice == "2":
                run_option2(drive_links, service_mgr)
            else:
                run_option3(drive_links, service_mgr)

            print(f"\n{SEP}")
            cont = input("  ↩️  Quay lại menu chính? (y/n, mặc định y): ").strip().lower()
            if cont not in ("", "y", "yes"):
                print("\n  👋 Thoát.\n")
                break

        else:
            print("\n  ⚠️  Lựa chọn không hợp lệ. Nhập 0, 1, 2 hoặc 3.")


if __name__ == "__main__":
    main()
